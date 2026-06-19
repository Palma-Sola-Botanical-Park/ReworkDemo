#!/usr/bin/env python3
"""
add_taxon_ids_to_plants.py
==========================
Add an **iNat Taxon ID** column to the PSBP plant signage sheet, so plants match
iNaturalist as solidly as the animals already do (no more name resolution at run
time in any downstream script).

It fills each ID from `taxon_cache.json` first (the cache the audit script already
wrote), and only calls the iNaturalist API for any botanical name not in the cache.
Run it after the audit and it's basically instant and offline.

The new column is inserted right after "Botanical Name". All other tabs
(Export Summary, Placements) and the original file are left untouched — output
goes to a NEW file.

Setup
-----
    pip3 install requests openpyxl

Run (from the folder with the sheet + taxon_cache.json)
------------------------------------------------------
    python3 add_taxon_ids_to_plants.py

Options
-------
    --plants PATH    input sheet (default: PSBP_Master_Plant_Signage-2026Jun18.xlsx)
    --out PATH       output sheet (default: <input>_withTaxonID.xlsx)
    --cache PATH     taxon cache json (default: taxon_cache.json)
    --no-api         only use the cache; never call the network (blanks stay blank)
"""

import argparse
import json
import os
import re
import sys
import time

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency. Run:  pip3 install requests openpyxl")

API_BASE      = "https://api.inaturalist.org/v1"
REQUEST_PAUSE = 1.0
USER_AGENT    = "PSBP-taxon-id-filler/1.0 (Palma Sola Botanical Park)"
SHEET_NAME    = "PSBP_Plants"
SCI_HEADER    = "Botanical Name"
RANK_HEADER   = "Sign Level"
NEW_HEADER    = "iNat Taxon ID"


def normalize(s):
    return re.sub(r"\s+", " ", str(s).strip().lower())


# Trailing " spp", " spp.", " sp", " sp." -> strip to bare genus.
SPP_RE = re.compile(r"\s+(?:spp|sp)\.?$", re.IGNORECASE)


def normalize_botanical(name):
    """Return (clean_name, changed). Collapses 'Canna spp.' -> 'Canna'."""
    s = re.sub(r"\s+", " ", str(name).strip())
    clean = SPP_RE.sub("", s).strip()
    return clean, (clean != s)


def get_json(session, url, params=None):
    for attempt in range(1, 4):
        try:
            r = session.get(url, params=params, timeout=60)
            if r.status_code == 200:
                time.sleep(REQUEST_PAUSE)
                return r.json()
        except Exception:
            pass
        time.sleep(2 * attempt)
    return None


def resolve_taxon_id(session, name, rank_hint, cache):
    key = normalize(name)
    if key in cache:
        return cache[key]
    if session is None:
        return None
    params = {"q": name, "is_active": "true", "per_page": 10}
    if rank_hint:
        params["rank"] = rank_hint
    data = get_json(session, f"{API_BASE}/taxa", params=params)
    if (not data or not data.get("results")) and rank_hint:
        data = get_json(session, f"{API_BASE}/taxa",
                        params={"q": name, "is_active": "true", "per_page": 10})
    tid = None
    if data and data.get("results"):
        results = data["results"]
        exact = [r for r in results if normalize(r.get("name", "")) == key]
        pick = exact[0] if exact else results[0]
        tid = pick.get("id")
    cache[key] = tid
    return tid


def header_map(ws):
    """Return {header_text: column_index} from row 1."""
    out = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val is not None and str(val).strip():
            out[str(val).strip()] = col
    return out


def main():
    ap = argparse.ArgumentParser(description="Add iNat Taxon ID column to plant sheet.")
    ap.add_argument("--plants", default="PSBP_Master_Plant_Signage-2026Jun18.xlsx")
    ap.add_argument("--out", default=None)
    ap.add_argument("--cache", default="taxon_cache.json")
    ap.add_argument("--no-api", action="store_true")
    args = ap.parse_args()

    if not os.path.isfile(args.plants):
        sys.exit(f"Input sheet not found: {args.plants}")
    out_path = args.out or os.path.splitext(args.plants)[0] + "_withTaxonID.xlsx"

    cache = {}
    if os.path.exists(args.cache):
        try:
            cache = json.load(open(args.cache))
            print(f"Loaded {len(cache)} cached name->id entries from {args.cache}.")
        except Exception:
            print(f"Could not read {args.cache}; starting with an empty cache.")
    elif args.no_api:
        print(f"No cache at {args.cache} and --no-api set; IDs will be blank.")

    session = None
    if not args.no_api:
        import requests
        session = requests.Session()
        session.headers.update({"User-Agent": USER_AGENT})

    wb = openpyxl.load_workbook(args.plants)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"Tab '{SHEET_NAME}' not found. Tabs present: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    hmap = header_map(ws)
    if SCI_HEADER not in hmap:
        sys.exit(f"Column '{SCI_HEADER}' not found. Headers: {list(hmap)}")
    sci_col = hmap[SCI_HEADER]
    rank_col = hmap.get(RANK_HEADER)

    # Insert (or reuse) the taxon-id column right after Botanical Name.
    if NEW_HEADER in hmap:
        id_col = hmap[NEW_HEADER]
        print(f"'{NEW_HEADER}' column already exists; refilling it.")
    else:
        id_col = sci_col + 1
        ws.insert_cols(id_col)
        ws.cell(row=1, column=id_col, value=NEW_HEADER)
        # inserting shifted everything at/after id_col right by one;
        # sci_col is to the left so it's unchanged, but re-read rank_col if needed.
        if rank_col is not None and rank_col >= id_col:
            rank_col += 1

    from_cache = from_api = blank = 0
    unresolved = []
    normalized = []   # (row, original, clean)
    rank_fixed = []   # (row, clean) where we set Sign Level -> Genus
    for r in range(2, ws.max_row + 1):
        raw = ws.cell(row=r, column=sci_col).value
        if raw is None or not str(raw).strip():
            continue

        # Normalize "Canna spp." -> "Canna" and write the clean name back.
        name, changed = normalize_botanical(raw)
        if changed:
            ws.cell(row=r, column=sci_col, value=name)
            normalized.append((r, str(raw).strip(), name))

        # Rank hint: bare-genus normalization forces genus; else read Sign Level.
        if changed:
            rank_hint = "genus"
            if rank_col:
                cur = str(ws.cell(row=r, column=rank_col).value or "").strip().lower()
                if cur != "genus":
                    ws.cell(row=r, column=rank_col, value="Genus")
                    rank_fixed.append((r, name))
        else:
            rank_hint = None
            if rank_col:
                rv = str(ws.cell(row=r, column=rank_col).value or "").strip().lower()
                rank_hint = rv if rv in ("species", "genus") else None

        key = normalize(name)
        in_cache = key in cache
        tid = resolve_taxon_id(session, name, rank_hint, cache)
        if tid:
            ws.cell(row=r, column=id_col, value=tid)
            from_cache += int(in_cache)
            from_api += int(not in_cache)
        else:
            blank += 1
            unresolved.append((r, str(name).strip()))

    # Save back the (possibly grown) cache and the new workbook.
    try:
        json.dump(cache, open(args.cache, "w"), indent=0)
    except Exception:
        pass
    wb.save(out_path)

    print(f"\nFilled {from_cache} from cache, {from_api} from API, {blank} left blank.")
    if normalized:
        print(f"\nNormalized {len(normalized)} name(s) to bare genus:")
        for r, old, new in normalized:
            print(f"  row {r}: '{old}' -> '{new}'")
    if rank_fixed:
        print(f"Set Sign Level -> Genus on {len(rank_fixed)} row(s): "
              + ", ".join(n for _, n in rank_fixed))
    if unresolved:
        print("\nUnresolved (check the botanical name / synonym):")
        for r, n in unresolved:
            print(f"  row {r}: {n}")
    print(f"\nWrote: {out_path}")
    print("Original file untouched. Open the new one, eyeball the IDs, then rename "
          "it over the original when you're happy.")


if __name__ == "__main__":
    main()
